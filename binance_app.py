#!/usr/bin/env python3
"""
Binance Portfolio Reporter — Desktop App
Pulls transactions from Binance and generates an XLSX report with:
  - All Transactions (trades, deposits, withdrawals)
  - Holdings (current balances with avg cost)
  - PNL Summary (profit/loss per coin and total)
"""

import hmac
import hashlib
import json
import os
import sys
import threading
import time
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime, timedelta
from pathlib import Path
import urllib.parse

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

# ---------------------------------------------------------------------------
# Binance API helpers (self-contained, no .env dependency)
# ---------------------------------------------------------------------------

RECV_WINDOW_MS = 60_000
API_RATE_LIMIT_SLEEP = 0.2
BASE_TS_MS = int(datetime(2020, 1, 1).timestamp() * 1000)
WINDOW_DAYS = 90


def sign_params(params: dict, api_secret: str) -> str:
    query = urllib.parse.urlencode(params, doseq=True)
    signature = hmac.new(
        api_secret.encode(), query.encode(), hashlib.sha256
    ).hexdigest()
    return f"{query}&signature={signature}"


def send_signed_request(method, path, params, api_key, api_secret, base_url):
    headers = {"X-MBX-APIKEY": api_key}
    params["timestamp"] = int(time.time() * 1000)
    params.setdefault("recvWindow", RECV_WINDOW_MS)
    qs = sign_params(params, api_secret)
    url = f"{base_url}{path}?{qs}"
    resp = requests.request(method, url, headers=headers, timeout=15)
    resp.raise_for_status()
    return resp.json()


def get_spot_price(symbol, base_url):
    try:
        resp = requests.get(
            f"{base_url}/api/v3/ticker/price",
            params={"symbol": symbol},
            timeout=10,
        )
        resp.raise_for_status()
        return float(resp.json()["price"])
    except Exception:
        return None


def to_iso(ts):
    if ts is None:
        return ""
    if isinstance(ts, (int, float)):
        return datetime.fromtimestamp(ts / 1000).isoformat()
    if isinstance(ts, datetime):
        return ts.isoformat()
    if isinstance(ts, str):
        try:
            return datetime.fromisoformat(ts.replace("Z", "")).isoformat()
        except ValueError:
            return ts
    return str(ts)


def to_ms(ts):
    if ts is None:
        return None
    if isinstance(ts, (int, float)):
        return int(ts)
    if isinstance(ts, str):
        ts = ts.replace("Z", "")
        try:
            return int(ts)
        except ValueError:
            try:
                return int(datetime.fromisoformat(ts).timestamp() * 1000)
            except ValueError:
                return None
    if isinstance(ts, datetime):
        return int(ts.timestamp() * 1000)
    return None


# ---------------------------------------------------------------------------
# Data fetching
# ---------------------------------------------------------------------------

def get_all_trades(symbol, api_key, api_secret, base_url, progress_cb=None):
    trades = []
    from_id = None
    while True:
        params = {"symbol": symbol.upper(), "limit": 1000}
        if from_id is not None:
            params["fromId"] = from_id
        data = send_signed_request("GET", "/api/v3/myTrades", params, api_key, api_secret, base_url)
        if not data:
            break
        trades.extend(data)
        if progress_cb:
            progress_cb(f"  {symbol}: fetched {len(trades)} trades...")
        if len(data) < 1000:
            break
        from_id = data[-1]["id"] + 1
        time.sleep(API_RATE_LIMIT_SLEEP)
    return trades


def fetch_deposits_windowed(api_key, api_secret, base_url, progress_cb=None):
    window_ms = int(timedelta(days=WINDOW_DAYS).total_seconds() * 1000)
    end_ms = int(time.time() * 1000)
    cur = BASE_TS_MS
    results = []
    while cur <= end_ms:
        win_end = min(cur + window_ms - 1, end_ms)
        while True:
            try:
                params = {"limit": 1000, "startTime": cur, "endTime": win_end}
                batch = send_signed_request("GET", "/sapi/v1/capital/deposit/hisrec", params, api_key, api_secret, base_url)
            except requests.HTTPError:
                return results
            if not batch:
                break
            results.extend(batch)
            max_ts = max((to_ms(d.get("insertTime")) or 0 for d in batch), default=0)
            if len(batch) < 1000 or max_ts <= 0:
                break
            cur = max_ts + 1
        cur = win_end + 1 if cur <= win_end else cur + 1
    if progress_cb:
        progress_cb(f"  Fetched {len(results)} deposits")
    return results


def fetch_withdrawals_windowed(api_key, api_secret, base_url, progress_cb=None):
    window_ms = int(timedelta(days=WINDOW_DAYS).total_seconds() * 1000)
    end_ms = int(time.time() * 1000)
    cur = BASE_TS_MS
    results = []
    while cur <= end_ms:
        win_end = min(cur + window_ms - 1, end_ms)
        while True:
            try:
                params = {"limit": 1000, "startTime": cur, "endTime": win_end}
                batch = send_signed_request("GET", "/sapi/v1/capital/withdraw/history", params, api_key, api_secret, base_url)
            except requests.HTTPError:
                return results
            if not batch:
                break
            results.extend(batch)
            max_ts = max((to_ms(w.get("applyTime")) or 0 for w in batch), default=0)
            if len(batch) < 1000 or max_ts <= 0:
                break
            cur = max_ts + 1
        cur = win_end + 1 if cur <= win_end else cur + 1
    if progress_cb:
        progress_cb(f"  Fetched {len(results)} withdrawals")
    return results


def get_fx_rate(ticker_symbol, default):
    try:
        import yfinance as yf
        ticker = yf.Ticker(ticker_symbol)
        hist = ticker.history(period="1d", interval="1d")
        if not hist.empty and "Close" in hist:
            rate = float(hist["Close"].dropna().iloc[-1])
            if rate > 0:
                return rate
    except Exception:
        pass
    return default


# ---------------------------------------------------------------------------
# Portfolio computation
# ---------------------------------------------------------------------------

def compute_holdings(trades_by_symbol: dict, deposits: list, withdrawals: list):
    """
    Compute per-coin holdings from trades, deposits and withdrawals.
    Returns dict: coin -> {buy_qty, buy_quote, sell_qty, sell_quote, deposit_qty, withdraw_qty}
    """
    holdings = {}

    def ensure(coin):
        if coin not in holdings:
            holdings[coin] = {
                "buy_qty": 0.0, "buy_quote": 0.0,
                "sell_qty": 0.0, "sell_quote": 0.0,
                "deposit_qty": 0.0, "withdraw_qty": 0.0,
            }

    for symbol, trades in trades_by_symbol.items():
        # Derive base coin from symbol (e.g. BTCUSDC -> BTC, ETHUSDT -> ETH)
        base_coin = None
        for quote in ("USDC", "USDT", "BUSD", "USD", "FDUSD"):
            if symbol.upper().endswith(quote):
                base_coin = symbol.upper()[: -len(quote)]
                break
        if not base_coin:
            base_coin = symbol[:3]  # fallback

        ensure(base_coin)
        for t in trades:
            try:
                qty = float(t.get("qty", 0))
                quote_qty = float(t.get("quoteQty", 0))
            except (TypeError, ValueError):
                continue
            if t.get("isBuyer"):
                holdings[base_coin]["buy_qty"] += qty
                holdings[base_coin]["buy_quote"] += quote_qty
            else:
                holdings[base_coin]["sell_qty"] += qty
                holdings[base_coin]["sell_quote"] += quote_qty

    for d in deposits:
        coin = d.get("coin", "")
        if not coin:
            continue
        ensure(coin)
        try:
            holdings[coin]["deposit_qty"] += float(d.get("amount", 0))
        except (TypeError, ValueError):
            pass

    for w in withdrawals:
        coin = w.get("coin", "")
        if not coin:
            continue
        ensure(coin)
        try:
            holdings[coin]["withdraw_qty"] += float(w.get("amount", 0))
        except (TypeError, ValueError):
            pass

    return holdings


# ---------------------------------------------------------------------------
# XLSX generation
# ---------------------------------------------------------------------------

# Style constants
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
GREEN_FONT = Font(name="Calibri", color="006100")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FONT = Font(name="Calibri", color="9C0006")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ZEBRA_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="595959")


def style_header(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def style_data_cell(cell, row_idx):
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center")
    if row_idx % 2 == 0:
        cell.fill = ZEBRA_FILL


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def generate_xlsx(
    trades_by_symbol: dict,
    deposits: list,
    withdrawals: list,
    live_prices: dict,
    fx_rates: dict,
    output_path: str,
):
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Transactions ----
    ws_tx = wb.active
    ws_tx.title = "Transactions"
    ws_tx.sheet_properties.tabColor = "2F5496"

    # Title
    ws_tx.merge_cells("A1:I1")
    title_cell = ws_tx["A1"]
    title_cell.value = "Binance Transaction History"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(vertical="center")
    ws_tx.row_dimensions[1].height = 30

    ws_tx.merge_cells("A2:I2")
    sub_cell = ws_tx["A2"]
    sub_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    sub_cell.font = SUBTITLE_FONT

    tx_headers = ["Date/Time", "Type", "Symbol/Coin", "Side", "Price (USD)", "Quantity", "Total (USD)", "Commission", "Status"]
    header_row = 4
    for col, h in enumerate(tx_headers, 1):
        ws_tx.cell(row=header_row, column=col, value=h)
    style_header(ws_tx, header_row, len(tx_headers))
    ws_tx.auto_filter.ref = f"A{header_row}:I{header_row}"
    ws_tx.freeze_panes = f"A{header_row + 1}"

    # Build rows sorted by time
    all_rows = []

    for symbol, trades in trades_by_symbol.items():
        for t in trades:
            time_str = to_iso(t.get("time"))
            side = "BUY" if t.get("isBuyer") else "SELL"
            try:
                price = float(t.get("price", 0))
                qty = float(t.get("qty", 0))
                quote = float(t.get("quoteQty", 0))
                comm = float(t.get("commission", 0))
            except (TypeError, ValueError):
                price = qty = quote = comm = 0
            comm_str = f"{comm:.8f} {t.get('commissionAsset', '')}" if comm else ""
            all_rows.append((time_str, "Trade", symbol, side, price, qty, quote, comm_str, "Filled"))

    for d in deposits:
        time_str = to_iso(d.get("insertTime"))
        try:
            amount = float(d.get("amount", 0))
        except (TypeError, ValueError):
            amount = 0
        status_map = {0: "Pending", 1: "Success", 6: "Credited"}
        status = status_map.get(d.get("status"), str(d.get("status", "")))
        all_rows.append((time_str, "Deposit", d.get("coin", ""), "", "", amount, "", "", status))

    for w in withdrawals:
        time_str = to_iso(w.get("applyTime"))
        try:
            amount = float(w.get("amount", 0))
        except (TypeError, ValueError):
            amount = 0
        status_map = {0: "Email Sent", 1: "Cancelled", 2: "Awaiting Approval",
                      3: "Rejected", 4: "Processing", 5: "Failure", 6: "Completed"}
        status = status_map.get(w.get("status"), str(w.get("status", "")))
        all_rows.append((time_str, "Withdrawal", w.get("coin", ""), "", "", amount, "", "", status))

    # Sort by time
    def sort_key(r):
        try:
            return datetime.fromisoformat(r[0])
        except Exception:
            return datetime.min

    all_rows.sort(key=sort_key)

    for i, row in enumerate(all_rows):
        r = header_row + 1 + i
        for col, val in enumerate(row, 1):
            cell = ws_tx.cell(row=r, column=col, value=val)
            style_data_cell(cell, i)
            # Format number columns
            if col == 5 and isinstance(val, (int, float)) and val:
                cell.number_format = '#,##0.00'
            elif col == 6 and isinstance(val, (int, float)) and val:
                cell.number_format = '#,##0.00000000'
            elif col == 7 and isinstance(val, (int, float)) and val:
                cell.number_format = '#,##0.00'

    auto_width(ws_tx)

    # ---- Sheet 2: Holdings ----
    ws_hold = wb.create_sheet("Holdings")
    ws_hold.sheet_properties.tabColor = "548235"

    ws_hold.merge_cells("A1:J1")
    title_cell = ws_hold["A1"]
    title_cell.value = "Current Holdings"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(vertical="center")
    ws_hold.row_dimensions[1].height = 30

    ws_hold.merge_cells("A2:J2")
    sub_cell = ws_hold["A2"]
    sub_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    sub_cell.font = SUBTITLE_FONT

    holdings = compute_holdings(trades_by_symbol, deposits, withdrawals)

    hold_headers = [
        "Coin", "Bought", "Sold", "Deposited", "Withdrawn",
        "Net Holdings", "Avg Buy Price (USD)", "Current Price (USD)",
        "Market Value (USD)", "Unrealised P&L (USD)",
    ]
    header_row = 4
    for col, h in enumerate(hold_headers, 1):
        ws_hold.cell(row=header_row, column=col, value=h)
    style_header(ws_hold, header_row, len(hold_headers))
    ws_hold.freeze_panes = f"A{header_row + 1}"

    row_idx = 0
    total_market = 0.0
    total_cost = 0.0

    # Sort coins: major coins first, then alphabetical
    major = ["BTC", "ETH", "BNB", "SOL"]
    sorted_coins = sorted(
        holdings.keys(),
        key=lambda c: (0 if c in major else 1, major.index(c) if c in major else 0, c),
    )

    for coin in sorted_coins:
        h = holdings[coin]
        net = h["buy_qty"] - h["sell_qty"] + h["deposit_qty"] - h["withdraw_qty"]
        if abs(net) < 1e-12 and h["buy_qty"] == 0:
            continue  # Skip coins with zero activity

        avg_buy = (h["buy_quote"] / h["buy_qty"]) if h["buy_qty"] > 0 else 0
        current_price = live_prices.get(coin, 0)
        market_val = net * current_price if current_price else 0
        cost_basis = h["buy_quote"] - h["sell_quote"]  # net cost
        pnl = market_val - cost_basis if current_price else 0

        if current_price:
            total_market += market_val
            total_cost += cost_basis

        r = header_row + 1 + row_idx
        values = [
            coin,
            h["buy_qty"],
            h["sell_qty"],
            h["deposit_qty"],
            h["withdraw_qty"],
            net,
            avg_buy,
            current_price if current_price else "N/A",
            market_val if current_price else "N/A",
            pnl if current_price else "N/A",
        ]
        for col, val in enumerate(values, 1):
            cell = ws_hold.cell(row=r, column=col, value=val)
            style_data_cell(cell, row_idx)
            # Number formatting
            if col in (2, 3, 4, 5, 6) and isinstance(val, float):
                cell.number_format = '#,##0.00000000'
            elif col in (7, 8, 9) and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
            elif col == 10 and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
                if val > 0:
                    cell.font = GREEN_FONT
                    cell.fill = GREEN_FILL
                elif val < 0:
                    cell.font = RED_FONT
                    cell.fill = RED_FILL

        row_idx += 1

    # Totals row
    if row_idx > 0:
        r = header_row + 1 + row_idx + 1
        ws_hold.cell(row=r, column=1, value="TOTAL").font = Font(bold=True, size=11)
        total_mv_cell = ws_hold.cell(row=r, column=9, value=total_market)
        total_mv_cell.font = Font(bold=True, size=11)
        total_mv_cell.number_format = '#,##0.00'
        total_pnl = total_market - total_cost
        total_pnl_cell = ws_hold.cell(row=r, column=10, value=total_pnl)
        total_pnl_cell.font = Font(bold=True, size=11, color="006100" if total_pnl >= 0 else "9C0006")
        total_pnl_cell.number_format = '#,##0.00'

    auto_width(ws_hold)

    # ---- Sheet 3: PNL Summary ----
    ws_pnl = wb.create_sheet("PNL Summary")
    ws_pnl.sheet_properties.tabColor = "BF8F00"

    ws_pnl.merge_cells("A1:H1")
    title_cell = ws_pnl["A1"]
    title_cell.value = "Profit & Loss Summary"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(vertical="center")
    ws_pnl.row_dimensions[1].height = 30

    ws_pnl.merge_cells("A2:H2")
    sub_cell = ws_pnl["A2"]
    sub_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    sub_cell.font = SUBTITLE_FONT

    pnl_headers = [
        "Coin", "Total Bought (USD)", "Total Sold (USD)", "Net Cost (USD)",
        "Current Holdings", "Market Value (USD)", "P&L (USD)", "P&L (%)",
    ]
    header_row = 4
    for col, h in enumerate(pnl_headers, 1):
        ws_pnl.cell(row=header_row, column=col, value=h)
    style_header(ws_pnl, header_row, len(pnl_headers))
    ws_pnl.freeze_panes = f"A{header_row + 1}"

    row_idx = 0
    grand_cost = 0.0
    grand_market = 0.0
    grand_sold = 0.0

    for coin in sorted_coins:
        h = holdings[coin]
        net_qty = h["buy_qty"] - h["sell_qty"] + h["deposit_qty"] - h["withdraw_qty"]
        if h["buy_quote"] == 0 and h["sell_quote"] == 0:
            continue  # No trading activity

        current_price = live_prices.get(coin, 0)
        market_val = net_qty * current_price if current_price else 0
        net_cost = h["buy_quote"] - h["sell_quote"]
        pnl = market_val - net_cost if current_price else 0
        pnl_pct = (pnl / h["buy_quote"] * 100) if h["buy_quote"] > 0 and current_price else 0

        if current_price:
            grand_market += market_val
        grand_cost += net_cost
        grand_sold += h["sell_quote"]

        r = header_row + 1 + row_idx
        values = [
            coin,
            h["buy_quote"],
            h["sell_quote"],
            net_cost,
            net_qty,
            market_val if current_price else "N/A",
            pnl if current_price else "N/A",
            pnl_pct if current_price else "N/A",
        ]
        for col, val in enumerate(values, 1):
            cell = ws_pnl.cell(row=r, column=col, value=val)
            style_data_cell(cell, row_idx)
            if col in (2, 3, 4, 6) and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
            elif col == 5 and isinstance(val, float):
                cell.number_format = '#,##0.00000000'
            elif col == 7 and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
                if val > 0:
                    cell.font = GREEN_FONT
                    cell.fill = GREEN_FILL
                elif val < 0:
                    cell.font = RED_FONT
                    cell.fill = RED_FILL
            elif col == 8 and isinstance(val, (int, float)):
                cell.number_format = '0.0"%"'
                if val > 0:
                    cell.font = GREEN_FONT
                elif val < 0:
                    cell.font = RED_FONT

        row_idx += 1

    # Grand totals
    if row_idx > 0:
        r = header_row + 1 + row_idx + 1
        ws_pnl.cell(row=r, column=1, value="TOTAL").font = Font(bold=True, size=11)
        for col, val in [(4, grand_cost), (6, grand_market)]:
            c = ws_pnl.cell(row=r, column=col, value=val)
            c.font = Font(bold=True, size=11)
            c.number_format = '#,##0.00'

        grand_pnl = grand_market - grand_cost
        grand_pnl_pct = (grand_pnl / (grand_cost + grand_sold) * 100) if (grand_cost + grand_sold) > 0 else 0

        pnl_cell = ws_pnl.cell(row=r, column=7, value=grand_pnl)
        pnl_cell.font = Font(bold=True, size=11, color="006100" if grand_pnl >= 0 else "9C0006")
        pnl_cell.number_format = '#,##0.00'

        pct_cell = ws_pnl.cell(row=r, column=8, value=grand_pnl_pct)
        pct_cell.font = Font(bold=True, size=11, color="006100" if grand_pnl_pct >= 0 else "9C0006")
        pct_cell.number_format = '0.0"%"'

        # FX conversions
        r += 2
        ws_pnl.cell(row=r, column=1, value="Currency Conversions").font = SUBTITLE_FONT
        r += 1
        usd_sgd = fx_rates.get("USD_SGD", 1.33)
        usd_cny = fx_rates.get("USD_CNY", 7.10)

        for label, rate in [("USD", 1.0), ("SGD", usd_sgd), ("CNY", usd_cny)]:
            ws_pnl.cell(row=r, column=1, value=f"Market Value ({label})").font = Font(bold=True)
            mv_cell = ws_pnl.cell(row=r, column=2, value=grand_market * rate)
            mv_cell.number_format = '#,##0.00'

            ws_pnl.cell(row=r, column=4, value=f"P&L ({label})").font = Font(bold=True)
            pnl_c = ws_pnl.cell(row=r, column=5, value=grand_pnl * rate)
            pnl_c.number_format = '#,##0.00'
            pnl_c.font = GREEN_FONT if grand_pnl >= 0 else RED_FONT
            r += 1

        r += 1
        ws_pnl.cell(row=r, column=1, value="FX Rates").font = SUBTITLE_FONT
        r += 1
        ws_pnl.cell(row=r, column=1, value=f"USD/SGD = {usd_sgd:.4f}    USD/CNY = {usd_cny:.4f}")

    auto_width(ws_pnl)

    wb.save(output_path)
    return len(all_rows), len(holdings)


# ---------------------------------------------------------------------------
# Config persistence (save API key to ~/Library/Application Support/ or ~/.config/)
# ---------------------------------------------------------------------------

def get_config_dir():
    if sys.platform == "darwin":
        config_dir = Path.home() / "Library" / "Application Support" / "BinanceReporter"
    else:
        config_dir = Path.home() / ".config" / "binance-reporter"
    config_dir.mkdir(parents=True, exist_ok=True)
    return config_dir


def load_config():
    config_file = get_config_dir() / "config.json"
    if config_file.exists():
        try:
            with open(config_file, "r") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def save_config(config):
    config_file = get_config_dir() / "config.json"
    with open(config_file, "w") as f:
        json.dump(config, f, indent=2)


# ---------------------------------------------------------------------------
# GUI Application
# ---------------------------------------------------------------------------

class BinanceReporterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Binance Portfolio Reporter")
        self.root.resizable(True, True)

        # Set minimum size
        self.root.minsize(600, 520)

        # Try to set window size nicely
        self.root.geometry("650x560")

        # Load saved config
        self.config = load_config()

        self._build_ui()

    def _build_ui(self):
        # Main container with padding
        main = ttk.Frame(self.root, padding=20)
        main.pack(fill=tk.BOTH, expand=True)

        # Title
        title = ttk.Label(main, text="Binance Portfolio Reporter", font=("Helvetica", 18, "bold"))
        title.pack(pady=(0, 5))
        subtitle = ttk.Label(main, text="Generate XLSX reports from your Binance account", font=("Helvetica", 11))
        subtitle.pack(pady=(0, 15))

        # --- API Credentials Frame ---
        cred_frame = ttk.LabelFrame(main, text="API Credentials", padding=10)
        cred_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(cred_frame, text="API Key:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10), pady=2)
        self.api_key_var = tk.StringVar(value=self.config.get("api_key", ""))
        self.api_key_entry = ttk.Entry(cred_frame, textvariable=self.api_key_var, width=55)
        self.api_key_entry.grid(row=0, column=1, sticky=tk.EW, pady=2)

        ttk.Label(cred_frame, text="API Secret:").grid(row=1, column=0, sticky=tk.W, padx=(0, 10), pady=2)
        self.api_secret_var = tk.StringVar(value=self.config.get("api_secret", ""))
        self.api_secret_entry = ttk.Entry(cred_frame, textvariable=self.api_secret_var, width=55, show="*")
        self.api_secret_entry.grid(row=1, column=1, sticky=tk.EW, pady=2)

        # Show/hide secret toggle
        self.show_secret = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            cred_frame, text="Show", variable=self.show_secret,
            command=self._toggle_secret
        ).grid(row=1, column=2, padx=(5, 0))

        # Save credentials checkbox
        self.save_creds = tk.BooleanVar(value=bool(self.config.get("api_key")))
        ttk.Checkbutton(cred_frame, text="Remember credentials", variable=self.save_creds).grid(
            row=2, column=1, sticky=tk.W, pady=(5, 0)
        )

        cred_frame.columnconfigure(1, weight=1)

        # --- Settings Frame ---
        settings_frame = ttk.LabelFrame(main, text="Settings", padding=10)
        settings_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(settings_frame, text="Symbols:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10), pady=2)
        self.symbols_var = tk.StringVar(value=self.config.get("symbols", "BTCUSDC, ETHUSDC"))
        ttk.Entry(settings_frame, textvariable=self.symbols_var, width=45).grid(row=0, column=1, sticky=tk.EW, pady=2)
        ttk.Label(settings_frame, text="(comma-separated)", font=("Helvetica", 9)).grid(row=0, column=2, padx=(5, 0))

        ttk.Label(settings_frame, text="Base URL:").grid(row=1, column=0, sticky=tk.W, padx=(0, 10), pady=2)
        self.base_url_var = tk.StringVar(value=self.config.get("base_url", "https://api.binance.com"))
        ttk.Entry(settings_frame, textvariable=self.base_url_var, width=45).grid(row=1, column=1, sticky=tk.EW, pady=2)

        settings_frame.columnconfigure(1, weight=1)

        # --- Output Frame ---
        output_frame = ttk.LabelFrame(main, text="Output", padding=10)
        output_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(output_frame, text="Save to:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10), pady=2)

        default_path = str(Path.home() / "Desktop" / f"Binance_Report_{datetime.now().strftime('%Y%m%d')}.xlsx")
        self.output_var = tk.StringVar(value=default_path)
        ttk.Entry(output_frame, textvariable=self.output_var, width=45).grid(row=0, column=1, sticky=tk.EW, pady=2)
        ttk.Button(output_frame, text="Browse...", command=self._browse_output).grid(row=0, column=2, padx=(5, 0))

        output_frame.columnconfigure(1, weight=1)

        # --- Generate Button ---
        self.generate_btn = ttk.Button(
            main, text="Generate Report", command=self._start_generate,
            style="Accent.TButton"
        )
        self.generate_btn.pack(pady=(5, 10), ipadx=20, ipady=5)

        # --- Progress ---
        self.progress = ttk.Progressbar(main, mode="indeterminate", length=400)
        self.progress.pack(pady=(0, 5))

        self.status_var = tk.StringVar(value="Ready")
        self.status_label = ttk.Label(main, textvariable=self.status_var, font=("Helvetica", 10), wraplength=600)
        self.status_label.pack()

    def _toggle_secret(self):
        self.api_secret_entry.config(show="" if self.show_secret.get() else "*")

    def _browse_output(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialdir=str(Path.home() / "Desktop"),
            initialfile=f"Binance_Report_{datetime.now().strftime('%Y%m%d')}.xlsx",
        )
        if path:
            self.output_var.set(path)

    def _update_status(self, msg):
        self.status_var.set(msg)
        self.root.update_idletasks()

    def _start_generate(self):
        api_key = self.api_key_var.get().strip()
        api_secret = self.api_secret_var.get().strip()

        if not api_key or not api_secret:
            messagebox.showerror("Error", "Please enter your Binance API Key and Secret.")
            return

        output_path = self.output_var.get().strip()
        if not output_path:
            messagebox.showerror("Error", "Please select an output file path.")
            return

        # Save config if requested
        if self.save_creds.get():
            self.config["api_key"] = api_key
            self.config["api_secret"] = api_secret
        else:
            self.config.pop("api_key", None)
            self.config.pop("api_secret", None)
        self.config["symbols"] = self.symbols_var.get()
        self.config["base_url"] = self.base_url_var.get()
        save_config(self.config)

        # Disable button and start progress
        self.generate_btn.config(state=tk.DISABLED)
        self.progress.start(10)

        # Run in background thread
        thread = threading.Thread(
            target=self._generate_report,
            args=(api_key, api_secret),
            daemon=True,
        )
        thread.start()

    def _generate_report(self, api_key, api_secret):
        try:
            base_url = self.base_url_var.get().strip().rstrip("/")
            symbols_raw = self.symbols_var.get().strip()
            symbols = [s.strip().upper() for s in symbols_raw.split(",") if s.strip()]
            output_path = self.output_var.get().strip()

            # 1. Fetch trades
            trades_by_symbol = {}
            for sym in symbols:
                self.root.after(0, self._update_status, f"Fetching trades for {sym}...")
                try:
                    trades = get_all_trades(sym, api_key, api_secret, base_url,
                                            progress_cb=lambda msg: self.root.after(0, self._update_status, msg))
                    trades_by_symbol[sym] = trades
                except requests.HTTPError as e:
                    detail = ""
                    if e.response is not None:
                        try:
                            detail = e.response.json().get("msg", "")
                        except Exception:
                            detail = e.response.text[:200]
                    self.root.after(0, lambda d=detail, s=sym: messagebox.showwarning(
                        "API Warning", f"Failed to fetch trades for {s}: {d}"))

            # 2. Fetch deposits
            self.root.after(0, self._update_status, "Fetching deposit history...")
            deposits = fetch_deposits_windowed(api_key, api_secret, base_url,
                                                progress_cb=lambda msg: self.root.after(0, self._update_status, msg))

            # 3. Fetch withdrawals
            self.root.after(0, self._update_status, "Fetching withdrawal history...")
            withdrawals = fetch_withdrawals_windowed(api_key, api_secret, base_url,
                                                      progress_cb=lambda msg: self.root.after(0, self._update_status, msg))

            # 4. Fetch live prices
            self.root.after(0, self._update_status, "Fetching live prices...")
            live_prices = {}
            holdings = compute_holdings(trades_by_symbol, deposits, withdrawals)
            for coin in holdings:
                for quote in ("USDC", "USDT"):
                    price = get_spot_price(f"{coin}{quote}", base_url)
                    if price:
                        live_prices[coin] = price
                        break

            # 5. Fetch FX rates
            self.root.after(0, self._update_status, "Fetching FX rates...")
            fx_rates = {
                "USD_SGD": get_fx_rate("USDSGD=X", 1.33),
                "USD_CNY": get_fx_rate("USDCNY=X", 7.10),
            }

            # 6. Generate XLSX
            self.root.after(0, self._update_status, "Generating XLSX report...")
            tx_count, coin_count = generate_xlsx(
                trades_by_symbol, deposits, withdrawals,
                live_prices, fx_rates, output_path,
            )

            total_trades = sum(len(t) for t in trades_by_symbol.values())
            self.root.after(0, self._on_success, output_path, total_trades, len(deposits), len(withdrawals))

        except Exception as e:
            self.root.after(0, self._on_error, str(e))

    def _on_success(self, path, trades, deposits, withdrawals):
        self.progress.stop()
        self.generate_btn.config(state=tk.NORMAL)
        self._update_status(
            f"Done! {trades} trades, {deposits} deposits, {withdrawals} withdrawals"
        )
        messagebox.showinfo(
            "Report Generated",
            f"Report saved to:\n{path}\n\n"
            f"Trades: {trades}\n"
            f"Deposits: {deposits}\n"
            f"Withdrawals: {withdrawals}",
        )

    def _on_error(self, error_msg):
        self.progress.stop()
        self.generate_btn.config(state=tk.NORMAL)
        self._update_status(f"Error: {error_msg}")
        messagebox.showerror("Error", f"Failed to generate report:\n\n{error_msg}")


def main():
    root = tk.Tk()

    # macOS-specific tweaks
    if sys.platform == "darwin":
        try:
            root.tk.call("::tk::unsupported::MacWindowStyle", "style", root._w, "document", "closeBox resizable")
        except Exception:
            pass

    app = BinanceReporterApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
