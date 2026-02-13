#!/usr/bin/env python3
"""
Binance Portfolio Reporter — CLI
Reads .env for credentials, pulls all transactions, outputs XLSX to Desktop.

Usage:
    python binance_report.py
    python binance_report.py --symbols BTCUSDT,ETHUSDT
    python binance_report.py --output ~/Downloads/report.xlsx
"""

import hmac
import hashlib
import json
import os
import sys
import time
import argparse
from datetime import datetime, timedelta
from pathlib import Path
import urllib.parse

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---------------------------------------------------------------------------
# .env loader
# ---------------------------------------------------------------------------

def load_env(env_path=None):
    """Load .env file from given path or next to this script."""
    if env_path is None:
        env_path = Path(__file__).resolve().parent / ".env"
    else:
        env_path = Path(env_path)

    if not env_path.exists():
        return

    with open(env_path, "r") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" not in line:
                continue
            key, _, value = line.partition("=")
            key = key.strip()
            value = value.strip().strip("'\"")
            os.environ.setdefault(key, value)


# ---------------------------------------------------------------------------
# Binance API
# ---------------------------------------------------------------------------

RECV_WINDOW_MS = 60_000
RATE_LIMIT_SLEEP = 0.2
BASE_TS_MS = int(datetime(2020, 1, 1).timestamp() * 1000)
WINDOW_DAYS = 90


def sign_params(params, secret):
    query = urllib.parse.urlencode(params, doseq=True)
    sig = hmac.new(secret.encode(), query.encode(), hashlib.sha256).hexdigest()
    return f"{query}&signature={sig}"


def api_request(method, path, params, key, secret, base_url):
    headers = {"X-MBX-APIKEY": key}
    params["timestamp"] = int(time.time() * 1000)
    params.setdefault("recvWindow", RECV_WINDOW_MS)
    qs = sign_params(params, secret)
    resp = requests.request(method, f"{base_url}{path}?{qs}", headers=headers, timeout=15)
    resp.raise_for_status()
    return resp.json()


def get_spot_price(symbol, base_url):
    try:
        r = requests.get(f"{base_url}/api/v3/ticker/price", params={"symbol": symbol}, timeout=10)
        r.raise_for_status()
        return float(r.json()["price"])
    except Exception:
        return None


def to_iso(ts):
    if ts is None:
        return ""
    if isinstance(ts, (int, float)):
        return datetime.fromtimestamp(ts / 1000).isoformat()
    return str(ts)


def to_ms(ts):
    if isinstance(ts, (int, float)):
        return int(ts)
    if isinstance(ts, str):
        try:
            return int(datetime.fromisoformat(ts.replace("Z", "")).timestamp() * 1000)
        except Exception:
            return None
    return None


STABLECOINS = ("USDC", "USDT", "BUSD", "FDUSD", "USD")
QUOTE_CURRENCIES = ("USDC", "USDT")


def fetch_traded_symbols(key, secret, base_url):
    """Discover all trading pairs by checking account balances against known quotes."""
    account = api_request("GET", "/api/v3/account", {}, key, secret, base_url)
    coins = set()
    for b in account.get("balances", []):
        asset = b["asset"]
        free = float(b.get("free", 0))
        locked = float(b.get("locked", 0))
        if free > 0 or locked > 0:
            coins.add(asset)

    # Build symbol pairs for non-stablecoin assets
    symbols = []
    for coin in sorted(coins):
        if coin in STABLECOINS:
            continue
        for quote in QUOTE_CURRENCIES:
            symbols.append(f"{coin}{quote}")
    return symbols


def fetch_account_balances(key, secret, base_url):
    """Return current spot balances from account endpoint (free + locked)."""
    account = api_request("GET", "/api/v3/account", {}, key, secret, base_url)
    balances = {}
    for b in account.get("balances", []):
        asset = b.get("asset")
        if not asset:
            continue
        try:
            free = float(b.get("free", 0))
            locked = float(b.get("locked", 0))
        except (TypeError, ValueError):
            continue
        total = free + locked
        if abs(total) > 1e-12:
            balances[asset] = total
    return balances


def fetch_all_trades(symbol, key, secret, base_url):
    trades, from_id = [], None
    while True:
        params = {"symbol": symbol.upper(), "limit": 1000}
        if from_id is not None:
            params["fromId"] = from_id
        data = api_request("GET", "/api/v3/myTrades", params, key, secret, base_url)
        if not data:
            break
        trades.extend(data)
        if len(data) < 1000:
            break
        from_id = data[-1]["id"] + 1
        time.sleep(RATE_LIMIT_SLEEP)
    return trades


def fetch_windowed(endpoint, time_field, key, secret, base_url):
    window_ms = int(timedelta(days=WINDOW_DAYS).total_seconds() * 1000)
    end_ms = int(time.time() * 1000)
    cur, results = BASE_TS_MS, []
    while cur <= end_ms:
        win_end = min(cur + window_ms - 1, end_ms)
        while True:
            try:
                batch = api_request("GET", endpoint, {"limit": 1000, "startTime": cur, "endTime": win_end}, key, secret, base_url)
            except requests.HTTPError:
                return results
            if not batch:
                break
            results.extend(batch)
            max_ts = max((to_ms(d.get(time_field)) or 0 for d in batch), default=0)
            if len(batch) < 1000 or max_ts <= 0:
                break
            cur = max_ts + 1
        cur = win_end + 1 if cur <= win_end else cur + 1
    return results


def get_fx_rate(ticker_symbol, default):
    try:
        import yfinance as yf
        t = yf.Ticker(ticker_symbol)
        hist = t.history(period="1d", interval="1d")
        if not hist.empty and "Close" in hist:
            rate = float(hist["Close"].dropna().iloc[-1])
            if rate > 0:
                return rate
    except Exception:
        pass
    return default


def load_manual_dca_plans(script_dir):
    json_path = script_dir / "manual_dca_plans.json"
    if not json_path.exists():
        return []
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return []


def load_manual_transfers(script_dir):
    json_path = script_dir / "manual_transfers.json"
    if not json_path.exists():
        return []
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return []


# ---------------------------------------------------------------------------
# Holdings computation
# ---------------------------------------------------------------------------

def compute_holdings(trades_by_symbol, deposits, withdrawals, manual_plans=None, manual_transfers=None):
    holdings = {}
    manual_plans = manual_plans or []
    manual_transfers = manual_transfers or []

    def ensure(coin):
        if coin not in holdings:
            holdings[coin] = {"buy_qty": 0.0, "buy_quote": 0.0, "sell_qty": 0.0, "sell_quote": 0.0, "deposit_qty": 0.0, "withdraw_qty": 0.0}

    for symbol, trades in trades_by_symbol.items():
        base_coin = None
        for quote in ("USDC", "USDT", "BUSD", "USD", "FDUSD"):
            if symbol.upper().endswith(quote):
                base_coin = symbol.upper()[:-len(quote)]
                break
        if not base_coin:
            base_coin = symbol[:3]

        quote_coin = symbol.upper()[len(base_coin):] if base_coin else None
        ensure(base_coin)
        if quote_coin:
            ensure(quote_coin)
        for t in trades:
            try:
                qty = float(t.get("qty", 0))
                quote_qty = float(t.get("quoteQty", 0))
            except (TypeError, ValueError):
                continue
            if t.get("isBuyer"):
                holdings[base_coin]["buy_qty"] += qty
                holdings[base_coin]["buy_quote"] += quote_qty
                if quote_coin:
                    holdings[quote_coin]["sell_qty"] += quote_qty
                    holdings[quote_coin]["sell_quote"] += quote_qty
            else:
                holdings[base_coin]["sell_qty"] += qty
                holdings[base_coin]["sell_quote"] += quote_qty
                if quote_coin:
                    holdings[quote_coin]["buy_qty"] += quote_qty
                    holdings[quote_coin]["buy_quote"] += quote_qty

    for d in deposits:
        coin = d.get("coin", "")
        if coin:
            ensure(coin)
            try:
                holdings[coin]["deposit_qty"] += float(d.get("amount", 0))
            except (TypeError, ValueError):
                pass

    for w in withdrawals:
        coin = w.get("coin", "")
        if coin:
            ensure(coin)
            try:
                holdings[coin]["withdraw_qty"] += float(w.get("amount", 0))
            except (TypeError, ValueError):
                pass

    for plan in manual_plans:
        for alloc in plan.get("allocations", []):
            coin = alloc.get("coin")
            if not coin:
                continue
            ensure(coin)
            try:
                amt = float(alloc.get("amount", 0))
                price = float(alloc.get("avg_price", 0))
            except (TypeError, ValueError):
                continue
            holdings[coin]["buy_qty"] += amt
            holdings[coin]["buy_quote"] += amt * price

    for tx in manual_transfers:
        coin = tx.get("coin")
        if not coin:
            continue
        ensure(coin)
        try:
            amt = float(tx.get("amount", 0))
            quote_amt = float(tx.get("quote", 0))
        except (TypeError, ValueError):
            continue
        holdings[coin]["buy_qty"] += amt
        holdings[coin]["buy_quote"] += quote_amt

    return holdings


# ---------------------------------------------------------------------------
# XLSX generation
# ---------------------------------------------------------------------------

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"),
)
GREEN_FONT = Font(name="Calibri", color="006100")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FONT = Font(name="Calibri", color="9C0006")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ZEBRA_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="595959")


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def style_cell(cell, row_idx):
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center")
    if row_idx % 2 == 0:
        cell.fill = ZEBRA_FILL


def auto_width(ws):
    for col in ws.columns:
        first = col[0]
        if hasattr(first, 'column_letter'):
            max_len = max((len(str(c.value or "")) for c in col if not isinstance(c, openpyxl.cell.cell.MergedCell)), default=0)
            ws.column_dimensions[first.column_letter].width = min(max_len + 4, 40)


def generate_xlsx(
    trades_by_symbol,
    deposits,
    withdrawals,
    live_prices,
    fx_rates,
    output_path,
    manual_plans=None,
    manual_transfers=None,
    current_balances=None,
):
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Transactions ----
    ws = wb.active
    ws.title = "Transactions"
    ws.sheet_properties.tabColor = "2F5496"

    ws.merge_cells("A1:I1")
    ws["A1"].value = "Binance Transaction History"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:I2")
    ws["A2"].value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = SUBTITLE_FONT

    headers = ["Date/Time", "Type", "Symbol/Coin", "Side", "Price (USD)", "Quantity", "Total (USD)", "Commission", "Status"]
    hr = 4
    for c, h in enumerate(headers, 1):
        ws.cell(row=hr, column=c, value=h)
    style_header(ws, hr, len(headers))
    ws.auto_filter.ref = f"A{hr}:I{hr}"
    ws.freeze_panes = f"A{hr + 1}"

    rows = []
    for symbol, trades in trades_by_symbol.items():
        for t in trades:
            try:
                price, qty, quote, comm = float(t.get("price", 0)), float(t.get("qty", 0)), float(t.get("quoteQty", 0)), float(t.get("commission", 0))
            except (TypeError, ValueError):
                price = qty = quote = comm = 0
            comm_str = f"{comm:.8f} {t.get('commissionAsset', '')}" if comm else ""
            rows.append((to_iso(t.get("time")), "Trade", symbol, "BUY" if t.get("isBuyer") else "SELL", price, qty, quote, comm_str, "Filled"))

    for d in deposits:
        try:
            amt = float(d.get("amount", 0))
        except (TypeError, ValueError):
            amt = 0
        status_map = {0: "Pending", 1: "Success", 6: "Credited"}
        rows.append((to_iso(d.get("insertTime")), "Deposit", d.get("coin", ""), "", "", amt, "", "", status_map.get(d.get("status"), str(d.get("status", "")))))

    for w in withdrawals:
        try:
            amt = float(w.get("amount", 0))
        except (TypeError, ValueError):
            amt = 0
        status_map = {0: "Email Sent", 1: "Cancelled", 2: "Awaiting Approval", 3: "Rejected", 4: "Processing", 5: "Failure", 6: "Completed"}
        rows.append((to_iso(w.get("applyTime")), "Withdrawal", w.get("coin", ""), "", "", amt, "", "", status_map.get(w.get("status"), str(w.get("status", "")))))

    rows.sort(key=lambda r: (datetime.fromisoformat(r[0]) if r[0] else datetime.min))

    for i, row in enumerate(rows):
        r = hr + 1 + i
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            style_cell(cell, i)
            if c == 5 and isinstance(val, float) and val:
                cell.number_format = '#,##0.00'
            elif c == 6 and isinstance(val, float) and val:
                cell.number_format = '#,##0.00000000'
            elif c == 7 and isinstance(val, float) and val:
                cell.number_format = '#,##0.00'
    auto_width(ws)

    # ---- Sheet 2: Holdings ----
    ws2 = wb.create_sheet("Holdings")
    ws2.sheet_properties.tabColor = "548235"
    ws2.merge_cells("A1:J1")
    ws2["A1"].value = "Current Holdings"
    ws2["A1"].font = TITLE_FONT
    ws2.row_dimensions[1].height = 30
    ws2.merge_cells("A2:J2")
    ws2["A2"].value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws2["A2"].font = SUBTITLE_FONT

    holdings = compute_holdings(trades_by_symbol, deposits, withdrawals, manual_plans, manual_transfers)
    h_headers = ["Coin", "Bought", "Sold", "Deposited", "Withdrawn", "Net Holdings", "Avg Buy Price (USD)", "Current Price (USD)", "Market Value (USD)", "Unrealised P&L (USD)"]
    hr2 = 4
    for c, h in enumerate(h_headers, 1):
        ws2.cell(row=hr2, column=c, value=h)
    style_header(ws2, hr2, len(h_headers))
    ws2.freeze_panes = f"A{hr2 + 1}"

    major = ["BTC", "ETH", "BNB", "SOL"]
    all_coins = set(holdings.keys())
    if current_balances:
        all_coins |= {c for c, v in current_balances.items() if abs(v) > 1e-12}
    sorted_coins = sorted(all_coins, key=lambda c: (0 if c in major else 1, major.index(c) if c in major else 0, c))

    ri, total_market, total_cost = 0, 0.0, 0.0
    for coin in sorted_coins:
        h = holdings.get(coin, {"buy_qty": 0.0, "buy_quote": 0.0, "sell_qty": 0.0, "sell_quote": 0.0, "deposit_qty": 0.0, "withdraw_qty": 0.0})
        inferred_net = h["buy_qty"] - h["sell_qty"] + h["deposit_qty"] - h["withdraw_qty"]
        net = current_balances.get(coin, inferred_net) if current_balances else inferred_net
        if abs(net) < 1e-12 and h["buy_qty"] == 0:
            continue
        avg_buy = (h["buy_quote"] / h["buy_qty"]) if h["buy_qty"] > 0 else 0
        cprice = live_prices.get(coin, 0)
        mval = net * cprice if cprice else 0
        cost = h["buy_quote"] - h["sell_quote"]
        pnl = mval - cost if cprice else 0
        if cprice:
            total_market += mval
            total_cost += cost

        r = hr2 + 1 + ri
        vals = [coin, h["buy_qty"], h["sell_qty"], h["deposit_qty"], h["withdraw_qty"], net, avg_buy, cprice or "N/A", mval if cprice else "N/A", pnl if cprice else "N/A"]
        for c, val in enumerate(vals, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            style_cell(cell, ri)
            if c in (2, 3, 4, 5, 6) and isinstance(val, float):
                cell.number_format = '#,##0.00000000'
            elif c in (7, 8, 9) and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
            elif c == 10 and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
                if val > 0:
                    cell.font, cell.fill = GREEN_FONT, GREEN_FILL
                elif val < 0:
                    cell.font, cell.fill = RED_FONT, RED_FILL
        ri += 1

    if ri > 0:
        r = hr2 + ri + 2
        ws2.cell(row=r, column=1, value="TOTAL").font = Font(bold=True, size=11)
        c = ws2.cell(row=r, column=9, value=total_market)
        c.font = Font(bold=True, size=11)
        c.number_format = '#,##0.00'
        tp = total_market - total_cost
        c = ws2.cell(row=r, column=10, value=tp)
        c.font = Font(bold=True, size=11, color="006100" if tp >= 0 else "9C0006")
        c.number_format = '#,##0.00'
    auto_width(ws2)

    # ---- Sheet 3: PNL Summary ----
    ws3 = wb.create_sheet("PNL Summary")
    ws3.sheet_properties.tabColor = "BF8F00"
    ws3.merge_cells("A1:H1")
    ws3["A1"].value = "Profit & Loss Summary"
    ws3["A1"].font = TITLE_FONT
    ws3.row_dimensions[1].height = 30
    ws3.merge_cells("A2:H2")
    ws3["A2"].value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws3["A2"].font = SUBTITLE_FONT

    p_headers = ["Coin", "Total Bought (USD)", "Total Sold (USD)", "Net Cost (USD)", "Current Holdings", "Market Value (USD)", "P&L (USD)", "P&L (%)"]
    hr3 = 4
    for c, h in enumerate(p_headers, 1):
        ws3.cell(row=hr3, column=c, value=h)
    style_header(ws3, hr3, len(p_headers))
    ws3.freeze_panes = f"A{hr3 + 1}"

    ri, g_cost, g_market, g_sold = 0, 0.0, 0.0, 0.0
    for coin in sorted_coins:
        h = holdings.get(coin, {"buy_qty": 0.0, "buy_quote": 0.0, "sell_qty": 0.0, "sell_quote": 0.0, "deposit_qty": 0.0, "withdraw_qty": 0.0})
        inferred_net_qty = h["buy_qty"] - h["sell_qty"] + h["deposit_qty"] - h["withdraw_qty"]
        net_qty = current_balances.get(coin, inferred_net_qty) if current_balances else inferred_net_qty
        if h["buy_quote"] == 0 and h["sell_quote"] == 0:
            continue
        cprice = live_prices.get(coin, 0)
        mval = net_qty * cprice if cprice else 0
        net_cost = h["buy_quote"] - h["sell_quote"]
        pnl = mval - net_cost if cprice else 0
        pnl_pct = (pnl / h["buy_quote"] * 100) if h["buy_quote"] > 0 and cprice else 0
        if cprice:
            g_market += mval
        g_cost += net_cost
        g_sold += h["sell_quote"]

        r = hr3 + 1 + ri
        vals = [coin, h["buy_quote"], h["sell_quote"], net_cost, net_qty, mval if cprice else "N/A", pnl if cprice else "N/A", pnl_pct if cprice else "N/A"]
        for c, val in enumerate(vals, 1):
            cell = ws3.cell(row=r, column=c, value=val)
            style_cell(cell, ri)
            if c in (2, 3, 4, 6) and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
            elif c == 5 and isinstance(val, float):
                cell.number_format = '#,##0.00000000'
            elif c == 7 and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
                if val > 0:
                    cell.font, cell.fill = GREEN_FONT, GREEN_FILL
                elif val < 0:
                    cell.font, cell.fill = RED_FONT, RED_FILL
            elif c == 8 and isinstance(val, (int, float)):
                cell.number_format = '0.0"%"'
                if val > 0:
                    cell.font = GREEN_FONT
                elif val < 0:
                    cell.font = RED_FONT
        ri += 1

    if ri > 0:
        r = hr3 + ri + 2
        ws3.cell(row=r, column=1, value="TOTAL").font = Font(bold=True, size=11)
        for col, val in [(4, g_cost), (6, g_market)]:
            c = ws3.cell(row=r, column=col, value=val)
            c.font = Font(bold=True, size=11)
            c.number_format = '#,##0.00'
        g_pnl = g_market - g_cost
        g_pnl_pct = (g_pnl / (g_cost + g_sold) * 100) if (g_cost + g_sold) > 0 else 0
        c = ws3.cell(row=r, column=7, value=g_pnl)
        c.font = Font(bold=True, size=11, color="006100" if g_pnl >= 0 else "9C0006")
        c.number_format = '#,##0.00'
        c = ws3.cell(row=r, column=8, value=g_pnl_pct)
        c.font = Font(bold=True, size=11, color="006100" if g_pnl_pct >= 0 else "9C0006")
        c.number_format = '0.0"%"'

        # FX conversions
        usd_sgd = fx_rates.get("USD_SGD", 1.33)
        usd_cny = fx_rates.get("USD_CNY", 7.10)
        r += 2
        ws3.cell(row=r, column=1, value="Currency Conversions").font = SUBTITLE_FONT
        r += 1
        for label, rate in [("USD", 1.0), ("SGD", usd_sgd), ("CNY", usd_cny)]:
            ws3.cell(row=r, column=1, value=f"Market Value ({label})").font = Font(bold=True)
            ws3.cell(row=r, column=2, value=g_market * rate).number_format = '#,##0.00'
            ws3.cell(row=r, column=4, value=f"P&L ({label})").font = Font(bold=True)
            pc = ws3.cell(row=r, column=5, value=g_pnl * rate)
            pc.number_format = '#,##0.00'
            pc.font = GREEN_FONT if g_pnl >= 0 else RED_FONT
            r += 1
        r += 1
        ws3.cell(row=r, column=1, value=f"FX Rates: USD/SGD={usd_sgd:.4f}  USD/CNY={usd_cny:.4f}")
    auto_width(ws3)

    wb.save(output_path)
    return len(rows)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Generate Binance portfolio XLSX report")
    parser.add_argument("--symbols", default=None, help="Comma-separated trading pairs (default: BTCUSDC,ETHUSDC)")
    parser.add_argument("--output", "-o", default=None, help="Output XLSX path (default: ~/Desktop/Binance_Report_YYYYMMDD.xlsx)")
    parser.add_argument("--env", default=None, help="Path to .env file (default: .env next to this script)")
    args = parser.parse_args()

    # Load .env
    load_env(args.env)

    api_key = os.environ.get("BINANCE_API_KEY", "")
    api_secret = os.environ.get("BINANCE_API_SECRET", "")
    base_url = os.environ.get("BINANCE_BASE_URL", "https://api.binance.com").rstrip("/")

    if not api_key or not api_secret:
        print("Error: BINANCE_API_KEY and BINANCE_API_SECRET not found.")
        print(f"Edit your .env file: {Path(__file__).resolve().parent / '.env'}")
        sys.exit(1)

    symbols_str = args.symbols or os.environ.get("SYMBOLS", "")
    if symbols_str:
        symbols = [s.strip().upper() for s in symbols_str.split(",") if s.strip()]
    else:
        print("Discovering traded symbols from account...", end=" ", flush=True)
        symbols = fetch_traded_symbols(api_key, api_secret, base_url)
        print(f"{len(symbols)} pairs")

    output = args.output or str(Path.home() / "Desktop" / f"Binance_Report_{datetime.now().strftime('%Y%m%d')}.xlsx")

    script_dir = Path(__file__).resolve().parent
    manual_plans = load_manual_dca_plans(script_dir)
    manual_transfers = load_manual_transfers(script_dir)

    print(f"Binance Portfolio Reporter")
    print(f"Symbols: {', '.join(symbols)}")
    if manual_plans:
        print(f"Manual DCA plans: {len(manual_plans)}")
    if manual_transfers:
        print(f"Manual transfers: {len(manual_transfers)}")
    print()

    # 1. Fetch trades
    trades_by_symbol = {}
    for sym in symbols:
        print(f"Fetching trades for {sym}...", end=" ", flush=True)
        try:
            trades = fetch_all_trades(sym, api_key, api_secret, base_url)
            trades_by_symbol[sym] = trades
            print(f"{len(trades)} trades")
        except requests.HTTPError as e:
            detail = ""
            if e.response is not None:
                try:
                    detail = e.response.json().get("msg", "")
                except Exception:
                    detail = e.response.text[:200]
            print(f"FAILED ({detail})")

    # 2. Fetch deposits & withdrawals
    print("Fetching deposits...", end=" ", flush=True)
    deposits = fetch_windowed("/sapi/v1/capital/deposit/hisrec", "insertTime", api_key, api_secret, base_url)
    print(f"{len(deposits)}")

    print("Fetching withdrawals...", end=" ", flush=True)
    withdrawals = fetch_windowed("/sapi/v1/capital/withdraw/history", "applyTime", api_key, api_secret, base_url)
    print(f"{len(withdrawals)}")

    # 3. Current balances
    print("Fetching current balances...", end=" ", flush=True)
    current_balances = fetch_account_balances(api_key, api_secret, base_url)
    print(f"{len(current_balances)} assets")

    # 4. Live prices
    print("Fetching live prices...", end=" ", flush=True)
    holdings = compute_holdings(trades_by_symbol, deposits, withdrawals, manual_plans, manual_transfers)
    coins_for_pricing = set(holdings.keys()) | set(current_balances.keys())
    live_prices = {}
    for coin in coins_for_pricing:
        for quote in ("USDC", "USDT"):
            price = get_spot_price(f"{coin}{quote}", base_url)
            if price:
                live_prices[coin] = price
                break
    print(f"{len(live_prices)} coins")

    # 5. FX rates
    print("Fetching FX rates...", end=" ", flush=True)
    fx_rates = {"USD_SGD": get_fx_rate("USDSGD=X", 1.33), "USD_CNY": get_fx_rate("USDCNY=X", 7.10)}
    print("done")

    # 6. Generate XLSX
    print()
    tx_count = generate_xlsx(
        trades_by_symbol,
        deposits,
        withdrawals,
        live_prices,
        fx_rates,
        output,
        manual_plans,
        manual_transfers,
        current_balances,
    )
    print(f"Report saved: {output}")
    print(f"  {tx_count} transactions across 3 sheets (Transactions, Holdings, PNL Summary)")


if __name__ == "__main__":
    main()
